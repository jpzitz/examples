# use openpyxl to open an excel sheet
# use python classes to sort attributes of the excel sheet


import openpyxl
        

# workbook class with pointer to the workbook
class virpyt_Workbook():

    # constructor/attribute __init__() that takes a workbookname
    # and then uses openpyxl to open that workbook and store the
    # workbook pointer in a local member.  
    def __init__(self, filename):
        self._workbook = openpyxl.load_workbook(filename)
    

    # .sheetnames returns a list of sheetnames
    @property
    def sheetnames(self):
        return self._workbook.sheetnames


    # returns a list of VirPyTSheet
    @property
    def sheets(self):
                #wraps openpyxl sheet objects using Sheet class
        return [Sheet(self._workbook[sheetname], sheetname)
                for sheetname in self.sheetnames]

    
    def save(self):
        self._workbook.save()
        
        

# sheet class with pointers to sheets in the workbook        
class Sheet():
    
    def __init__(self, sheet, name):
        self._sheet = sheet
        self._name = name
        self._table = {}    #{startcell : values}
        

    @property
    def name(self):
        return self._name

    #table objects
    @property
    def tables(self):
        self.find_tables()
        
        return list(self._table.values())

        
    def find_tables(self):

        totalrow = self._sheet.min_row
        totalcol = self._sheet.min_column
        
        #find first cell with values using min_row, min_column
        startcell = self._sheet.cell(row=self._sheet.min_row,
                                     column=self._sheet.min_column).coordinate

        
        while totalrow < self._sheet.max_row:
            # 0-based numrow & numcol
            numrow = 0
            numcol = 0
            table_data = []
            

            # scan until empty column is found
            for col in self._sheet.iter_cols(min_col=totalcol,
                                             min_row=totalrow):
                if col[0].value:    #header row should extend over whole table
                    numcol += 1
                else:
                    break
            
            
            for row in self._sheet.iter_rows(min_row=totalrow,
                                             min_col=totalcol,
                                             max_col=totalcol+numcol-1):
                #sometimes theres a gap where one column doesnt have data
                emptyrow = True
                for cell in row:
                    if cell.value:
                        emptyrow = False
                        
                if emptyrow:
                    break
                    
                else:
                    numrow += 1
                    table_data.append(row)
                    

            totalrow +=numrow -1
            totalcol +=numcol -1


            self._table[startcell] = Table(startcell, numcol, numrow, table_data)
            

            # find next table
            startcell, totalrow, totalcol = self.startcell(totalrow,
                                                           totalcol,
                                                           numrow,
                                                           numcol)
            

        



    def startcell(self, rowcoord, colcoord, numrow, numcol):

        # search vertically for next table
        for row in self._sheet.iter_rows():
                if not row[self._sheet.min_row].value:
                    rowcoord += 1
        
        '''
        # start seraching horizontally if maxrows reached
        if rowcoord == self._sheet.max_row:
            for col in self._sheet.iter_cols(min_col = colcoord):
                if not col[0].value:
                    colcoord += 1

            return self._sheet.cell(row=(rowcoord-numrow+1),
                                    column=colcoord).coordinate
        else:
            return self._sheet.cell(row=rowcoord,
                                    column=(colcoord-numcol+1)).coordinate
        '''

        startcell = self._sheet.cell(row=rowcoord,
                                    column=(colcoord-numcol+1))
        return startcell.coordinate, startcell.row , startcell.column        

    


        #scan til valued cell, use header as startcell
        
        
        
        

# table class probably to scan empty cells that bound the table
# or look for cell border formatting in the file
class Table():
    def __init__(self, startcell, numrow, numcol, table):
        
        #defines table object with starting cell and dimensions
        self._startcell = startcell
        self._numrow = numrow
        self._numcol = numcol
        self.table = table

        
    @property
    def header(self):
        header = []
        for cell in self.table[0]:
            header.append(cell.value)

        return header

    @property
    def rows(self):
        return [([cell.value for cell in row]) for row in self.table]

    @property
    def columns(self):
        columns = []
        idx = 0
        while idx < len(self.table[0]):
            columns.append([row[idx].value for row in self.table])
            idx+=1
            
        return columns
            
            
            
        
        
        

        

class Row():
    
    def __init__(self, row ):
        self.row = row

            
#class Cell():

              

if __name__ == '__main__':
    #workbookname = input(print("Input workbookname: "))
    
    wb = virpyt_Workbook('sample.xlsx')
    print(wb)        #address of openpyxl workbook object

    print(wb.sheets)        #list of sheet object addresses
    
    print(wb.sheetnames)    #list of names of worksheets

    for sheet in wb.sheets:     #prints each sheet title
        print("Found sheet named %s" % sheet.name)
        for table in sheet.tables:
            print("Found table: ", table._startcell,
                  table._numrow, table._numcol)
            print(table.header)
            #print(table.rows[0])    #rows[0] same as header method
            print(table.columns[0])  #prints first column of data in each table
